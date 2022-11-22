import os
from re import A
import sys
import glob
import csv
import openpyxl
import pygame as pg
import Components as cmp
from tkinter import filedialog as fd
import time

# Constants
aborted = [False]
# Opening file explorer to select document
file_path = fd.askopenfilename(title ='Select Finance Document', filetypes=[('Excel files', '*.xlsx')])
if(file_path == ""):
    print("Invalid Finance Document")
    exit()

#Check when the finances were last checked
wb = openpyxl.load_workbook(file_path)
sh = wb[wb.sheetnames[0]] 
lastchecked = sh['E4'].value

# Load in the transactions
filelist = fd.askopenfilename(title ='Select Transactions', filetypes=[('CSV files', '*.csv')])
if(filelist == ""):
    print("Invalid Transaction Document")
    exit()
transactions = []
with open(filelist, 'r') as file:
    csvreader = csv.reader(file)
    for row in csvreader:
        temp_transaction = ".".join(row).split(';')
        transaction = [temp_transaction[5],temp_transaction[8],temp_transaction[9],temp_transaction[14].replace('                                                                       ', ""), temp_transaction[17].replace("                                                                                                                                            ", "")]
        temp1 = str(transaction).split(",")
        temp2 = lastchecked.split(",")
        if temp1[0] == temp2[0] and temp1[3] == temp2[3] and temp1[4] == temp2[4]:
            break
        else:
            transactions.append(transaction)

transactions.pop(0)
if len(transactions) == 0:
    exit()
sh['E4'] = str(transactions[0])

#Create abrv_dict with abreviations for each sheet
abrv_dict  = {}
sheetnames = wb.sheetnames
sheetnames.remove('Algemeen')
sheetnames.remove('Sjabloon')
sheetnames.remove('Transacties')
for name in sheetnames:
    abreviation = wb[name]['E4'].value
    if abreviation != None:
        abrv_dict[abreviation] = name

#Find lenght of all sheets
emptyrow_dict = {}
for sheetname in sheetnames:
    sheet = wb[sheetname]
    current_row = 1
    while sheet.cell(current_row,1).value != None:
        current_row += 1
    emptyrow_dict[sheetname] = current_row

#Find lenght of transaction sheet
sheet = wb["Transacties"]
transactie_row = 1
while sheet.cell(transactie_row,1).value != None:
        transactie_row += 1

#Setting up the GUI to manually sort unidentified transaction





# pg.init()
# screen = pg.display.set_mode((1400,800))
# pg.display.set_caption("Automated Finances")
# screen.fill((255, 255, 255))
# pg.display.update()

# while(len(transactions) == 0):
#     screen.fill((255, 255, 255))
#     pg.display.update()

#Create a list of buttons to choose the right sheetname
# buttons = draw.Buttons(sheetnames, screen)

#Processing the transactions that have a valid abreviation
transactions = transactions[::-1]
for transaction_id in range(len(transactions)):
    #Looking for right sheetname to add
    discription = transactions[transaction_id][4].split(" ")
    if discription[0].upper() in abrv_dict.keys():
        transactions[transaction_id].append(abrv_dict[discription[0].upper()])
        transactions[transaction_id][4] = " ".join(discription[1:]).upper()
    else:
        transactions[transaction_id].append("")
        # ready = False
        # transaction_graphic = draw.Transaction(transaction, screen)
        # while not ready:
        #     for event in pg.event.get():
        #         if event.type == pg.QUIT:
        #             ready = True
        #             aborted = True
        #         elif event.type == pg.MOUSEBUTTONDOWN:
        #             x,y = pg.mouse.get_pos()
        #             sheetname = buttons.click(x, y)
        #             if sheetname != None:
        #                 ready = True 
        #     screen.fill((255, 255, 255))
        #     transaction_graphic.draw()
        #     buttons.draw()
        #     pg.display.update()
gui = cmp.GUI()
gui.run(sheetnames, transactions, aborted)
if aborted[0]:
    exit()
for transaction in transactions:
    #Addding to the sheet
    sheet = wb[transaction[5]]
    sheet.cell(emptyrow_dict[transaction[5]], 1).value = transaction[3]+ ": " + transaction[4]
    sheet.cell(emptyrow_dict[transaction[5]], 2).value = float(transaction[1])
    sheet.cell(emptyrow_dict[transaction[5]], 2).number_format = '_-€ * #,##0.00_-;-€ * #,##0.00_-;_-€ * "-"??_-;_-@_-'
    emptyrow_dict[transaction[5]] += 1

    sheet = wb["Transacties"]
    sheet.cell(transactie_row, 1).value = transaction[0]
    sheet.cell(transactie_row, 2).value = transaction[3]
    sheet.cell(transactie_row, 3).value = transaction[4]
    sheet.cell(transactie_row, 4).value = float(transaction[1])
    sheet.cell(transactie_row, 4).number_format = '_-€ * #,##0.00_-;-€ * #,##0.00_-;_-€ * "-"??_-;_-@_-'
    sheet.cell(transactie_row, 5).value = transaction[5]
    sheet.cell(transactie_row, 6).value = transaction[2]
    transactie_row += 1

#Closing files
wb.save(file_path)
wb.close()


        